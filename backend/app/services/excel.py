"""Master spreadsheet generator (openpyxl).

The People DB table is the source of truth; the .xlsx is regenerated from it
on demand. Also makes a stable photo thumbnail per person (for the People page
and for embedding in the sheet).
"""
import os

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, Side

from app.config import get_settings

settings = get_settings()

PHOTO_MAX_PX = 150
ROW_HEIGHT = 115
TITLE = "LIST PESERTA PELATIHAN"


def master_path() -> str:
    os.makedirs(settings.storage_output, exist_ok=True)
    return os.path.join(settings.storage_output, "master.xlsx")


def photos_dir() -> str:
    d = os.path.join(settings.storage_output, "photos")
    os.makedirs(d, exist_ok=True)
    return d


def make_thumbnail(photo_src: str, job_id: str) -> str | None:
    """Save a stable PNG thumbnail at output/photos/<job_id>.png.

    Handles images + PDF first page. Returns the saved path or None.
    """
    if not photo_src or not os.path.exists(photo_src):
        return None
    from PIL import Image as PILImage

    dest = os.path.join(photos_dir(), f"{job_id}.png")
    try:
        if photo_src.lower().endswith(".pdf"):
            import cv2
            from app.services import pdf_converter
            page = pdf_converter.to_images(photo_src)[0]
            cv2.imwrite(dest, page)
            img = PILImage.open(dest)
        else:
            img = PILImage.open(photo_src)
        img = img.convert("RGB")
        img.thumbnail((PHOTO_MAX_PX, PHOTO_MAX_PX))
        img.save(dest, "PNG")
        return dest
    except Exception:
        return None


# Columns B..K, mirroring Sample.xlsx (A is a left margin).
# (column_letter, person attr or "_no" for row number / None for photo, width)
COLUMNS = [
    ("B", "_no", 6),               # NO
    ("C", None, 22),               # SCAN KTP (photo)
    ("D", "nik", 20),              # NIK
    ("E", "full_name", 26),       # NAMA
    ("F", "jenis_pelatihan", 22), # JENIS PELATIHAN (manual)
    ("G", "place_of_birth", 18),  # LAHIR > TEMPAT
    ("H", "date_of_birth", 14),   # LAHIR > TANGGAL
    ("I", "company_name", 26),    # PERUSAHAAN > NAMA
    ("J", "company_address", 40), # PERUSAHAAN > ALAMAT
    ("K", "ket", 14),             # KET (manual)
]

_THIN = Side(style="thin", color="DDDDDD")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def generate_workbook(people) -> str:
    """Build master.xlsx fresh from Person rows, matching the Sample.xlsx layout."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Peserta"

    ws.column_dimensions["A"].width = 3
    for letter, _attr, width in COLUMNS:
        ws.column_dimensions[letter].width = width

    # Title (row 2), merged B..K
    ws["B2"] = TITLE
    ws["B2"].font = Font(bold=True, size=14)
    ws["B2"].alignment = _CENTER
    ws.merge_cells("B2:K2")

    # Grouped headers rows 4-5
    hdr = Font(bold=True)
    for col, label in {"B": "NO", "C": "SCAN KTP", "D": "NIK", "E": "NAMA",
                       "F": "JENIS PELATIHAN", "K": "KET"}.items():
        ws.merge_cells(f"{col}4:{col}5")
        c = ws[f"{col}4"]
        c.value, c.font, c.alignment = label, hdr, _CENTER
    for col, label in (("G", "LAHIR"), ("I", "PERUSAHAAN")):
        end = chr(ord(col) + 1)
        ws.merge_cells(f"{col}4:{end}4")
        c = ws[f"{col}4"]
        c.value, c.font, c.alignment = label, hdr, _CENTER
    for col, label in (("G", "TEMPAT"), ("H", "TANGGAL"), ("I", "NAMA"), ("J", "ALAMAT")):
        c = ws[f"{col}5"]
        c.value, c.font, c.alignment = label, hdr, _CENTER
    for letter, _a, _w in COLUMNS:
        for r in (4, 5):
            ws[f"{letter}{r}"].border = _BORDER

    # Data rows from row 6
    for i, p in enumerate(people):
        row = 6 + i
        ws.row_dimensions[row].height = ROW_HEIGHT
        for letter, attr, _w in COLUMNS:
            cell = ws[f"{letter}{row}"]
            cell.border = _BORDER
            cell.alignment = _LEFT if attr in ("company_address", "company_name") else _CENTER
            if attr == "_no":
                cell.value = i + 1
            elif attr is not None:
                cell.value = getattr(p, attr, None)
        if p.photo_path and os.path.exists(p.photo_path):
            try:
                ws.add_image(XLImage(p.photo_path), f"C{row}")
            except Exception:
                pass

    path = master_path()
    wb.save(path)
    return path
